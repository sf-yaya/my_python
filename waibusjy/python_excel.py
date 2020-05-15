"""
每行都要写注释，直接提交代码
使用openpyxl实现以下需求

使用excel 写入一组数据，姓名，身高，体重
计算是否为健康体重，如果是健康体重，则在旁边备注健康，并将姓名打印出来
健康体重计算公式：（身高cm-70）×60%
(可以做一部分优化)

"""

from openpyxl import Workbook, load_workbook


class Identity_information:

    def create_date(self):
        wb = Workbook()
        dest_filename = 'Identity_information.xlsx'
        ws1 = wb.active
        ws1.title = "information"

        # 写入一组数据，姓名，身高，体重
        ws1["A1"] = "名字"
        ws1["B1"] = "身高"
        ws1["C1"] = "体重"
        name = ["yaya", "yeye", "fenfen", "花花", "小丽"]
        # height = [100, 170, 160, 150, 165]
        # weight = [50, 60, 55, 40, 70]
        # 字典类型
        data_dict = {100: 51, 170: 60, 160: 51, 150: 42, 165: 72}
        list1 = [50, 60, 55, 40, 70, 20, 60]

        # 列表推导式,获取字典中的身高数据
        data_kes = [t for t in data_dict.keys()]
        # 根据name的长度，对相对应的列表插入数据
        for row in range(len(name)):
            ws1.cell(row=row + 2, column=1).value = name[row]
            ws1.cell(row=row + 2, column=2).value = data_kes[row]
            ws1.cell(row=row + 2, column=3).value = data_dict[data_kes[row]]

        # 创建第二个页签
        ws2 = wb.create_sheet(title="Pi1")
        for j in range(len(list1)):
            ws2.cell(row=j + 1, column=j + 1).value = list1[j]
        # 创建第二个页签
        ws3 = wb.create_sheet(title="test1")
        ws3["A1"] = "我在测试pythone写入数据功能11"
        wb.save(dest_filename)

    # 读取文件
    def get_info(self):
        # 获取excel的名称
        wb = load_workbook(filename='Identity_information.xlsx')
        # 获取excel的页签的名称
        sheet_info = wb['information']
        # 新增一列“备注”信息
        sheet_info["D1"] = "备注"
        # 如何获取excel列表中的长度？而不是输入一个固定数据
        # 如何设置column的数据，而不是输入一个固定的列
        for i in range(4):
            name = sheet_info.cell(row=i + 2, column=1).value
            height = sheet_info.cell(row=i + 2, column=2).value
            weight = sheet_info.cell(row=i + 2, column=3).value
            # 健康体重公式
            health = (height - 70) * 0.6
            if weight == health:
                print(name, "是健康体重", weight)
                # 新增一列备注健康体重
                sheet_info.cell(row=i + 2, column=4, value="健康体重")
        # 保存excel修改后的数据
        wb.save("Identity_information.xlsx")


# 实例化类
info = Identity_information()
# 调用类中的方法
info.create_date()
info.get_info()
