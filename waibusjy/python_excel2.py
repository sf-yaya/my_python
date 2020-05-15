"""
每行都要写注释，直接提交代码
使用openpyxl实现以下需求

使用excel 写入一组数据，姓名，身高，体重
计算是否为健康体重，如果是健康体重，则在旁边备注健康，并将姓名打印出来
健康体重计算公式：（身高cm-70）×60%
(可以做一部分优化)

"""

from openpyxl import Workbook, load_workbook


class Identity_information:
    def __init__(self):
        self.wb = Workbook()
        self.ws1 = self.wb.active
        self.ws1.title = "information"
        self.dest_filename = 'Identity_information.xlsx'
        # 写入一组数据，姓名，身高，体重
        self.ws1["A1"] = "名字"
        self.ws1["B1"] = "身高"
        self.ws1["C1"] = "体重"
        self.name = ["yaya", "yanyan", "fenfen", "花花", "小丽"]
        # self.height = [100, 170, 160, 150, 161]
        # self.weight = [50, 60, 55, 40, 70]
        # 字典类型
        self.data_dict = {170: 60, 175: 63, 160: 52, 150: 42, 162: 45}
        self.list1 = [50, 60, 55, 40, 70, 20, 60]

    def create_date(self):
        # 列表推导式,获取data_dict字典中的身高数据
        # print(self.data_dict.keys())

        # 如果身高有2个或者多个相同的人，字典keys索引越界
        data_kes = [t for t in self.data_dict.keys()]
        # 根据name的长度，对相对应的列表插入数据
        for row in range(len(self.name)):
            self.ws1.cell(row=row + 2, column=1).value = self.name[row]
            self.ws1.cell(row=row + 2, column=2).value = data_kes[row]
            self.ws1.cell(row=row + 2, column=3).value = self.data_dict[data_kes[row]]

        # 自增代码
        # 创建第二个页签
        # ws2 = self.wb.create_sheet(title="Pi")
        # for j in range(len(self.list1)):
        #     ws2.cell(row=j + 1, column=j + 1).value = self.list1[j]
        # # 创建第二个页签
        # ws3 = self.wb.create_sheet(title="test")
        # ws3["A1"] = "我在测试pythone写入数据功能"

        # 创建成功后进行保存
        self.wb.save(self.dest_filename)

    # 读取文件
    def get_info(self):
        # 获取excel的名称
        wb = load_workbook(filename=self.dest_filename)
        # 获取excel的页签的名称
        sheet_info = wb['information']
        # 新增一列“备注”信息
        sheet_info["D1"] = "备注"
        # 如何获取excel列表中的长度？而不是输入一个固定数据
        # 如何设置column的数据，而不是输入一个固定的列
        for i in range(4):
            name = sheet_info.cell(row=i + 2, column=1).value
            height = sheet_info.cell(row=i + 2, column=2).value
            weight = sheet_info.cell(row=i + 2, column=3).value
            # 健康体重公式
            health = (height - 70) * 0.6
            if weight == health:
                print(str(name) + "的体重是" + str(weight) + "kg，非常健康")
                # 新增一列备注健康体重
                sheet_info.cell(row=i + 2, column=4, value="健康体重")
        wb.save(self.dest_filename)


# 实例化类
info = Identity_information()
# 调用类中的方法
info.create_date()
info.get_info()
